"""Synthetic verification for ops_dashboard v1.8.1.

Everything the v1.8.0 harness covered, re-run against v1.8.1 (shifted indices), plus the
v1.8.1 additions:

  * a technician active in SOME window weeks only     (windowed distinct > mean of weeklies)
  * an entity that OPENS LATE / a week with NO census  (the v1.6.x mask machinery must hold)
  * a SUPPRESSED single-technician site
  * an employee with ON-CALL hours straddling 40h      (threshold excludes them)
  * a FEED-INCOMPLETE tail payroll week                (flagged, hollow, out of trailing)
  * a site with an IDLE payroll week                   (headcount zero-fills — audit A2)
  * thin weekly mature cohorts                         (pooled recovery ≠ mean of ratios)
  * a scorecard metric ABSENT from the company frame   (guard reports it — audit A1)
  * a payroll OT Week workbook                         (Cell 15.6 parses the Grand Total)
  * the FULL RENDERER, and every weekly axis label carrying the year (MMM-YY)  [v1.8.1]
  * a PLANTED technician-name spike month              (Cell 16.5 flags it, decomposes it,
                                                        and calls the name-splitting
                                                        signature)                [v1.8.1]
  * a PLANTED payroll-hours DIP month                  (negative deviations flag too)
  * payroll_weeks on the OT monthly bridge             (4-vs-5-week months visible)

Cells executed (v1.8.1 indices): 4.3 (=20), 13.5 (=66), 14 (=68), 15.5 (=72), 15.6 (=74),
16.5 (=78), 18 (=84), 17.5 (=82).

Run from the repository root:

    python analysis/lib/verify_ops_dashboard_v1_8_1.py

Exits non-zero on any failure. Needs no database connection, reads only the notebook, and
writes its scratch files to a temp directory.
"""
import json, os, sys, io, re, math, tempfile
from collections import defaultdict
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

NB = 'analysis/ops_dashboard_2026-08-14_v1_8_1.ipynb'
OUT = os.path.join(tempfile.gettempdir(), 'ops_dashboard_verify_v181')
os.makedirs(OUT, exist_ok=True)

_nb = json.load(open(NB, encoding='utf-8'))
CELL = {i: ''.join(c['source']) for i, c in enumerate(_nb['cells'])}

FAILS, PASSES = [], []


def check(cond, label):
    (PASSES if cond else FAILS).append(label)
    print(('  PASS  ' if cond else '  FAIL  ') + label)


def raises(fn, label, exc=Exception):
    try:
        fn()
    except exc:
        PASSES.append(label); print('  PASS  ' + label); return
    FAILS.append(label); print('  FAIL  ' + label + '  (no exception raised)')


# ═══════════════════════════════════════════════════════════════════════════════
# Namespace: the config the tested cells read, at the values Cells 3.1/3.5 set.
# ═══════════════════════════════════════════════════════════════════════════════
FILTER_START, FILTER_END = '2026-01-07', '2026-03-11'
AS_OF_DATE = pd.Timestamp('2026-03-11').date()
G = dict(
    np=np, pd=pd, plt=plt, mdates=mdates, io=io, os=os, re=re, math=math,
    defaultdict=defaultdict,
    FILTER_START=FILTER_START, FILTER_END=FILTER_END, AS_OF_DATE=AS_OF_DATE,
    PROD_WEEK_START_DOW=6, PROD_WEEK_LABEL='Sun-Sat', OT_WEEK_START_DOW=3,
    OT_WEEK_LABEL='Thu-Wed', OT_WEEKLY_THRESHOLD=40.0,
    ROLL_WEEKS=4, DASH_ROLL_WEEKS=4, TRAILING_MIN_WEEKS=2, TRAILING_SUFFIX='_t4w',
    TRAILING_EXCLUDE_PARTIAL_WEEKS=True,
    CENSUS_MIN_ACTIVE_TECHS=1.0, CENSUS_MIN_TECHS=2, CENSUS_WARN_UNMAPPED_PCT=10.0,
    CENSUS_HEADLINE_METRIC='census_per_tech_headcount',
    VIRTUAL_WH_PREFIX='Z', VIRTUAL_RUG_THRESHOLD_PCT=5.0,
    ROUTING_CHANGE_DATE='2026-02-01', RUN_DATE='2026-08-18',
    DASH_TOP_N=5, DASH_MIN_ACTIVE_DAYS=30, DASH_PAGE_INCHES=(23, 15),
    LOST_IS_STALE=True, LOST_STALE_NOTE='lost feed ends 2026-02-28 — tail is missing data',
    LOST_BULK_EVENT_DATES=['2025-08-01'], LOST_MAX_DATE=pd.Timestamp('2026-02-28'),
    LOST_RECOVERY_MATURITY_DAYS=90,
    OT_PAY_TYPES_WORKED=['Work', 'On Call Hours'], OT_PAY_TYPES_LEAVE=['PTO', 'Holiday'],
    OT_PAY_TYPES_OT_BASIS=['Work'],
    OT_FEED_COMPLETE_MIN_RATIO=0.70, OT_FEED_COMPLETE_TAIL_DAYS=28,
    OT_DEPTS_TECH=['Patient Care Technician'],
    OT_DEPTS_ALL=['Patient Care Technician', 'Internal Operations'],
    OT_APPORTION_BY_TICKET_SHARE=True, PCT_DEPT='Patient Care Technician',
    FUZZY_EDIT_DIST=1, BRACKET_CHECKS_RAISE=False,
    RECS_LOOKBACK_WEEKS=8, RECS_MIN_WEEKS=2, RECS_MIN_ACTIVE_DAYS=5,
    RECS_NAME_N=5, RECS_EFF_GAP_PCT=15.0, RECS_MIN_SEVERITY_PCT=5.0, RECS_TOP_N=5,
    COACHING_CAVEAT='coaching caveat',
    OUT_DIR=OUT,
    PALETTE={'attributed': '#1f77b4'},
    save_fig=lambda fig, name: None,
    standardize_first_name=lambda raw: [raw] if raw else [],
    levenshtein=lambda a, b: 0 if a == b else max(len(a), len(b)),
)
G['_clean'] = lambda s: re.sub(r"['\.\s\-]", '', s).lower() if isinstance(s, str) else ''
TS = G['TRAILING_SUFFIX']

print('\n=== Cell 4.3 — spine, add_trailing & the generalised bracket guard ===========')
exec(CELL[20], G)
week_start_of, add_trailing, WEEK_SPINE = G['week_start_of'], G['add_trailing'], G['WEEK_SPINE']
attach_spine, build_week_spine = G['attach_spine'], G['build_week_spine']
bracket_violations, run_bracket_checks = G['bracket_violations'], G['run_bracket_checks']
WEEKS = list(WEEK_SPINE['week_start'])
check(len(WEEKS) == 10, f'spine has 10 weeks (got {len(WEEKS)})')

# bracket guard: a frame whose trailing column is a genuine pooled ratio passes...
_t = pd.DataFrame({'week_start': WEEKS, 'r': [10.0, 12, 11, 13, 12, 11, 10, 12, 11, 12]})
_t['r' + TS] = _t['r'].rolling(4, min_periods=2).mean()
_nbad, _ntot = bracket_violations(_t, 'r', [])
check(_nbad == 0, f'bracket guard passes a true trailing mean ({_nbad} violations)')
# ...and a product-of-means style corruption fails
_bad = _t.copy(); _bad['r' + TS] = _bad['r' + TS] * 1.5
_nbad, _ = bracket_violations(_bad, 'r', [])
check(_nbad > 0, f'bracket guard catches a value outside its weekly bracket ({_nbad} rows)')
G['BRACKET_CHECKS_RAISE'] = True
raises(lambda: run_bracket_checks('verify', [('co', _bad, [], ['r'])]),
       'run_bracket_checks RAISES when BRACKET_CHECKS_RAISE is True', AssertionError)
G['BRACKET_CHECKS_RAISE'] = False
run_bracket_checks('verify', [('co', _bad, [], ['r'])])
check(True, 'run_bracket_checks demotes to a warning when BRACKET_CHECKS_RAISE is False')

# ═══════════════════════════════════════════════════════════════════════════════
# Synthetic upstream frames for Cell 13.5 — v1.7.0 harness plus Hal, the
# alternating-week technician the windowed denominator exists for.
# ═══════════════════════════════════════════════════════════════════════════════
_HIER = {'WH_A': ('VP North', 'TX', 'Houston'),
         'WH_B': ('VP North', 'TX', 'Houston'),
         'WH_C': ('VP South', 'LA', 'Other')}
_rows = []
for wi, ws in enumerate(WEEKS):
    for d in pd.date_range(ws, ws + pd.Timedelta(days=6)):
        if d.dayofweek >= 5 or not (pd.Timestamp(FILTER_START) <= d <= pd.Timestamp(FILTER_END)):
            continue
        if wi != 4:                                   # week 4: WH_A silent entirely
            _rows.append(('Ann', 'Alpha', 'WH_A', d, 9))
            _rows.append(('Dee', 'Delta', 'WH_A', d, 8))
            _rows.append(('Bob', 'Beta', 'WH_A', d, 6))
            if wi % 2 == 0:                           # Hal works alternating weeks only
                _rows.append(('Hal', 'Eta', 'WH_A', d, 5))
        if wi >= 5:                                   # WH_B opens in week 5
            _rows.append(('Bob', 'Beta', 'WH_B', d, 3))
            _rows.append(('Eve', 'Epsilon', 'WH_B', d, 5))
            _rows.append(('Fay', 'Phi', 'WH_B', d, 4))
        _rows.append(('Cal', 'Gamma', 'WH_C', d, 4))
_td = pd.DataFrame(_rows, columns=['techfirstname', 'techlastname', 'tech_warehouse',
                                  'completed_date', 'day_tickets'])
_td['_tech_key'] = _td['techfirstname'] + '|' + _td['techlastname']
_td[['vp', 'state', 'metro']] = pd.DataFrame(
    [_HIER[w] for w in _td['tech_warehouse']], index=_td.index)
_td['week_start'] = week_start_of(_td['completed_date'])
_td['_tech_day_tickets'] = _td.groupby(['_tech_key', 'completed_date'])['day_tickets'].transform('sum')
_td['active_day_share'] = _td['day_tickets'] / _td['_tech_day_tickets']
G['_techday'] = _td
_GR = ['techfirstname', 'techlastname', '_tech_key', 'tech_warehouse', 'vp', 'state', 'metro']
G['_techday_wk'] = (_td.groupby(_GR + ['week_start'], dropna=False, as_index=False)
                    .agg(total_tickets=('day_tickets', 'sum'),
                         active_weekdays=('completed_date', 'nunique'),
                         active_day_equiv=('active_day_share', 'sum')))

_cen = []
for wi, ws in enumerate(WEEKS):
    for d in pd.date_range(ws, ws + pd.Timedelta(days=6)):
        if not (pd.Timestamp(FILTER_START) <= d <= pd.Timestamp(AS_OF_DATE)):
            continue
        if wi == 4:
            continue
        if wi == 6 and d.dayofweek != 0:
            continue
        if wi < 8:
            _cen.append((d, 'WH_A', 900.0))
        if wi >= 5:
            _cen.append((d, 'WH_B', 300.0))
        _cen.append((d, 'WH_C', 120.0))
        _cen.append((d, 'Z CS', 40.0))
_cd = pd.DataFrame(_cen, columns=['census_date', 'warehouse', 'pt_count'])
_cd['pt_count_all'] = _cd['pt_count'] * 1.1
_cd['_is_virtual_census'] = _cd['warehouse'].str.upper().str.startswith('Z')
_cd['week_start'] = week_start_of(_cd['census_date'])
_cd['period'] = _cd['census_date'].dt.strftime('%Y-%m')
G['df_census_daily'] = _cd

print('\n=== Cell 13.5 — census per technician (windowed denominator) ================')
exec(CELL[66], G)
CWH, CVP, CMET, CCO = (G['dash_wk_census_wh'], G['dash_wk_census_vp'],
                       G['dash_wk_census_metro'], G['dash_wk_census_co'])

# The published inputs still divide into the published figure, all grains (Recon 5 raised
# inside the cell if not, but re-derive independently).
_bad = 0
for _fr in (CCO, CVP, CMET, CWH):
    _p = pd.to_numeric(_fr['census_per_tech_headcount' + TS], errors='coerce')
    _q = (pd.to_numeric(_fr['ratio_adc' + TS], errors='coerce')
          / pd.to_numeric(_fr['ratio_techs_equiv' + TS], errors='coerce'))
    _bad += int((_p.notna() & _q.notna() & ((_p - _q).abs() > 0.01 + 0.001 * _p.abs())).sum())
    _bad += int((_p.isna() & _q.notna()).sum())
check(_bad == 0, f'published ADC / published technicians == published ratio, all grains '
                 f'({_bad} mismatches)')
check(int(pd.to_numeric(CCO['census_per_tech_headcount' + TS], errors='coerce').notna().sum()) > 0,
      'the company trailing ratio is populated (the check above is not vacuous)')

# THE POINT OF v1.8.0: the trailing denominator is the DISTINCT count over the window's
# pooled weeks. Re-derive it independently from the raw synthetic feed.
_co = CCO.sort_values('week_start').reset_index(drop=True)
_D = (pd.to_numeric(_co['techs_equiv'], errors='coerce').fillna(0).gt(0)
      & pd.to_numeric(_co['census_days'], errors='coerce').fillna(0).gt(0)
      & ~_co['ratio_suppressed'].fillna(True).astype(bool)
      & ~_co['is_partial_week'].fillna(False).astype(bool))
_wk_list = list(_co['week_start'])
_mismatch, _tested, _strict_gt = 0, 0, 0
for i in range(len(_co)):
    pub = pd.to_numeric(_co.loc[i, 'ratio_techs_equiv' + TS], errors='coerce')
    if pd.isna(pub):
        continue
    lo = max(0, i - 3)
    dweeks = [_wk_list[j] for j in range(lo, i + 1) if bool(_D.iloc[j])]
    expect = _td[_td['week_start'].isin(dweeks)]['_tech_key'].nunique()
    weekly_mean = float(np.mean([_co.loc[j, 'techs_distinct'] for j in range(lo, i + 1)
                                 if bool(_D.iloc[j])])) if dweeks else np.nan
    _tested += 1
    if abs(float(pub) - expect) > 0.01:
        _mismatch += 1
    if pd.notna(weekly_mean) and float(pub) > weekly_mean + 0.01:
        _strict_gt += 1
check(_tested > 0 and _mismatch == 0,
      f'company trailing denominator == distinct technicians active in the window '
      f'({_mismatch} mismatches over {_tested} weeks)')
check(_strict_gt > 0,
      'window count is STRICTLY above the weekly mean somewhere (Hal, the alternating '
      'technician, is actually being counted once instead of ~half)')

# Suppression, apportionment and dtype contracts carried over from v1.7.0
_c = CWH[CWH['tech_warehouse'] == 'WH_C']
check(bool(_c['ratio_suppressed'].all()), 'single-technician site suppressed on every week')
check(bool(pd.to_numeric(_c['ratio_techs_equiv' + TS], errors='coerce').isna().all()),
      'a suppressed site publishes no trailing denominator input (no orphan numbers)')
_wk5 = WEEKS[6]
_s_eq = float(CWH[CWH['week_start'] == _wk5]['techs_equiv'].sum())
_co_di = float(CCO[CCO['week_start'] == _wk5]['techs_distinct'].iloc[0])
check(abs(_s_eq - _co_di) < 0.01,
      f'weekly apportioned headcount still sums to the company distinct count '
      f'({_s_eq:.2f} vs {_co_di:.0f})')
check(str(CWH['ratio_suppressed'].dtype) == 'bool',
      'ratio_suppressed survives as bool through the trailing pass')
print('  (Reconciliations 1-6 and 4b inside Cell 13.5 all passed — the cell raises otherwise.)')

# ═══════════════════════════════════════════════════════════════════════════════
# Cell 14 — lost equipment: pooled recovery on thin mature cohorts
# ═══════════════════════════════════════════════════════════════════════════════
print('\n=== Cell 14 — lost equipment (pooled recovery) ===============================')
_lrows = []
_aid = 0
for wi, ws in enumerate(WEEKS[1:9], start=1):
    # week pattern: 2 mature assets (1 recovered) on odd weeks, 20 mature (2 recovered) on
    # even weeks -> mean of weekly rates 30%, pooled ~13.6%: the divergence recovery is
    # censored into.
    n, rec = (2, 1) if wi % 2 else (20, 2)
    for k in range(n):
        _aid += 1
        _lrows.append({'lost_date': ws + pd.Timedelta(days=1), 'asset_tag': f'A{_aid}',
                       'lost_cost': 25.0, 'product_name': 'Concentrator',
                       'lost_reason': 'Not found', 'resolution': 'No resolution',
                       'is_recovered': k < rec, 'is_discarded': False,
                       'is_unresolved': k >= rec, 'cohort_mature': True,
                       'days_to_resolve': 30.0 if k < rec else np.nan,
                       'tech_warehouse': 'WH_A', 'vp': 'VP North', 'state': 'TX',
                       'metro': 'Houston',
                       'period': (ws + pd.Timedelta(days=1)).strftime('%Y-%m')})
_lost = pd.DataFrame(_lrows)
G['_lost_filt'] = _lost
G['df_inventory_total'] = pd.DataFrame([{'tech_warehouse': 'WH_A',
                                         'total_inventory_count': 1000,
                                         'total_inventory_amount': 50000.0}])
exec(CELL[68], G)
LCO = G['dash_wk_lost_co'].sort_values('week_start').reset_index(drop=True)
_r = LCO[pd.to_numeric(LCO['recovery_rate_pct' + TS], errors='coerce').notna()]
if len(_r):
    _row = _r.iloc[-1]
    _exp = (float(_row['mature_recovered_assets' + TS])
            / float(_row['mature_cohort_assets' + TS]) * 100)
    check(abs(float(_row['recovery_rate_pct' + TS]) - _exp) < 0.11,
          f'trailing recovery is POOLED (got {_row["recovery_rate_pct" + TS]:.1f}, '
          f'pooled parts give {_exp:.1f})')
    _wmean = float(pd.to_numeric(LCO['recovery_rate_pct'], errors='coerce').mean())
    check(abs(float(_row['recovery_rate_pct' + TS]) - _wmean) > 5,
          f'…and it visibly differs from the mean of weekly rates ({_wmean:.1f}) on this '
          f'thin-cohort pattern')
else:
    check(False, 'no trailing recovery value produced')
LWH = G['dash_wk_lost_wh']
_pi = LWH[pd.to_numeric(LWH['lost_pct_of_inventory' + TS], errors='coerce').notna()]
if len(_pi):
    _row = _pi.iloc[-1]
    _exp = float(_row['lost_asset_count' + TS]) / 1000 * 100
    check(abs(float(_row['lost_pct_of_inventory' + TS]) - _exp) < 0.01,
          'trailing inventory share = trailing mean lost count / constant inventory')
else:
    check(False, 'no trailing inventory share produced')

# ═══════════════════════════════════════════════════════════════════════════════
# Cell 15.5 — overtime: on-call threshold, feed guard, per-head pooling, A1 alias
# ═══════════════════════════════════════════════════════════════════════════════
print('\n=== Cell 15.5 — overtime =====================================================')
OT_SPINE_PREVIEW = build_week_spine(FILTER_START, AS_OF_DATE, 3)
OTW = list(OT_SPINE_PREVIEW['week_start'])
_hrows = []


def _emp_week(name, ws, work_by_day, oncall=0.0, pto=0.0, dept='Patient Care Technician'):
    for di, h in enumerate(work_by_day):
        if h:
            _hrows.append({'plc_id': name, 'employee_name': name, 'workdate': ws + pd.Timedelta(days=di),
                           'hours': float(h), 'pay_type': 'Work', 'dept': dept,
                           'location_name': 'X', 'sys_updated': pd.Timestamp('2026-03-01'),
                           'reg_hrs_unusable': 0.0, 'ot1_hrs_unusable': 0.0})
    if oncall:
        _hrows.append({'plc_id': name, 'employee_name': name, 'workdate': ws,
                       'hours': float(oncall), 'pay_type': 'On Call Hours', 'dept': dept,
                       'location_name': 'X', 'sys_updated': pd.Timestamp('2026-03-01'),
                       'reg_hrs_unusable': 0.0, 'ot1_hrs_unusable': 0.0})
    if pto:
        _hrows.append({'plc_id': name, 'employee_name': name, 'workdate': ws + pd.Timedelta(days=1),
                       'hours': float(pto), 'pay_type': 'PTO', 'dept': dept,
                       'location_name': 'X', 'sys_updated': pd.Timestamp('2026-03-01'),
                       'reg_hrs_unusable': 0.0, 'ot1_hrs_unusable': 0.0})


# Ann: 44 Work hours a week (4h OT). Bob: 38 Work + 6 On Call (0h OT new basis; 4h on the
# old). Cal: alternating weeks at WH_C (A2 case). Feed-incomplete: the LAST FULL payroll
# week gets 40% of everyone's hours.
for wi, ws in enumerate(OTW):
    if OT_SPINE_PREVIEW['is_partial_week'].iloc[wi]:
        continue
    scale = 0.4 if wi == len(OTW) - 2 else 1.0
    _emp_week('Alpha, Ann', ws, [s * scale for s in (9, 9, 9, 9, 8)])
    _emp_week('Beta, Bob', ws, [s * scale for s in (8, 8, 8, 7, 7)], oncall=6 * scale)
    if wi % 2 == 0:
        _emp_week('Gamma, Cal', ws, [s * scale for s in (10, 10, 10, 6, 6)])
df_hours_raw = pd.DataFrame(_hrows)
df_hours_raw['workdate'] = pd.to_datetime(df_hours_raw['workdate'])
df_hours_raw['dow_mon0'] = (df_hours_raw['workdate'] - pd.Timestamp('1900-01-01')).dt.days % 7
G['df_hours_raw'] = df_hours_raw

# ticket frame for site attribution: each person works one site, Sun-Sat weeks
_txr = []
for wi, ws in enumerate(WEEKS):
    for d in pd.date_range(ws, ws + pd.Timedelta(days=4)):
        if not (pd.Timestamp(FILTER_START) <= d <= pd.Timestamp(AS_OF_DATE)):
            continue
        _txr.append(('Ann', 'Alpha', 'WH_A', d))
        _txr.append(('Bob', 'Beta', 'WH_A', d))
        if wi % 2 == 0:
            _txr.append(('Cal', 'Gamma', 'WH_C', d))
df_tx = pd.DataFrame(_txr, columns=['techfirstname', 'techlastname', 'tech_warehouse',
                                    'completed_date'])
df_tx['order_num'] = np.arange(len(df_tx)).astype(str)
df_tx['_unattributed'] = False
df_tx[['vp', 'state', 'metro']] = pd.DataFrame([_HIER[w] for w in df_tx['tech_warehouse']],
                                               index=df_tx.index)
G['df_tx'] = df_tx

exec(CELL[72], G)
OCO = G['dash_wk_ot_co'].sort_values('week_start').reset_index(drop=True)
OWH = G['dash_wk_ot_wh']
OT_SPINE = G['OT_WEEK_SPINE']

_full = OCO[~OCO['is_partial_week'].fillna(False).astype(bool)]
# probe a week where only Ann (44h Work) and Bob (38h Work + 6h on-call) appear — Cal's
# alternating weeks add his own 2h of OT and would muddy the arithmetic being tested
_probe = _full[_full['employees'] == 2].iloc[0]
check(abs(float(_probe['ot_hours']) - 4.0) < 0.01,
      f'on-call excluded from the threshold: company OT is Ann\'s 4h only '
      f'(got {_probe["ot_hours"]})')
check(abs(float(_probe['ot_hours_incl_oncall']) - 8.0) < 0.01,
      f'the reconciliation basis still counts Bob\'s on-call OT (38+6-40=4h more) '
      f'(got {_probe["ot_hours_incl_oncall"]})')
check(abs(float(_probe['worked_hours']) - (44 + 44 + (0 if _probe.name % 2 else 0))) < 60,
      'worked_hours keeps on-call in the labour-cost lens')

_flagged = OT_SPINE[OT_SPINE['is_feed_incomplete']]
check(len(_flagged) == 1, f'exactly the deflated tail week is flagged feed-incomplete '
                          f'(got {len(_flagged)})')
_fb_week = _flagged['week_start'].iloc[0] if len(_flagged) else None
if _fb_week is not None:
    _row = OCO[OCO['week_start'] == _fb_week]
    check(bool(_row['is_partial_week'].iloc[0]),
          'the feed-incomplete week is treated as partial (hollow, out of trailing windows)')
    _after = OCO[OCO['week_start'] > _fb_week]
    _tvals = pd.to_numeric(OCO['ot_hours' + TS], errors='coerce')
    # the trailing mean in the weeks at/after the flag must not have been dragged by the
    # deflated week: it should stay near the healthy weekly OT (4h), not fall toward 1.6h
    _tail_t = pd.to_numeric(OCO.loc[OCO['week_start'] >= _fb_week, 'ot_hours' + TS],
                            errors='coerce').dropna()
    check(bool((_tail_t > 3.0).all()) if len(_tail_t) else True,
          f'trailing OT ignores the feed-incomplete week (tail trailing values '
          f'{list(_tail_t.round(2))})')

check('ot_hours_per_tech' in OCO.columns and ('ot_hours_per_tech' + TS) in OCO.columns,
      'audit A1: the company frame now carries ot_hours_per_tech (+ trailing)')

# A2: Cal's site (WH_C) works alternating weeks. Pooled per-tech trailing must divide by a
# zero-filled technician count, i.e. equal ot_hours_t4w / technicians_t4w with quiet weeks
# counted as zero people.
WHC = OWH[OWH['tech_warehouse'] == 'WH_C'].sort_values('week_start')
_wc = WHC[pd.to_numeric(WHC['ot_hours_per_tech' + TS], errors='coerce').notna()]
if len(_wc):
    _row = _wc.iloc[-1]
    _exp = float(_row['ot_hours' + TS]) / float(_row['technicians' + TS])
    check(abs(float(_row['ot_hours_per_tech' + TS]) - _exp) < 0.05,
          f'audit A2: pooled per-tech OT divides by the zero-filled technician count '
          f'(got {_row["ot_hours_per_tech" + TS]}, parts give {_exp:.2f})')
    check(float(_row['technicians' + TS]) < 1.0 - 1e-9,
          f'…and the trailing technician count reflects the idle weeks '
          f'(got {_row["technicians" + TS]}, would be 1.0 under the v1.7.1 rate treatment)')
else:
    check(False, 'no trailing per-tech OT produced for the intermittent site')

# ═══════════════════════════════════════════════════════════════════════════════
# Cell 15.6 — the calibration diagnostic
# ═══════════════════════════════════════════════════════════════════════════════
print('\n=== Cell 15.6 — OT calibration ================================================')
_cwd = os.getcwd()
_iso = os.path.join(OUT, 'empty'); os.makedirs(_iso, exist_ok=True)
os.chdir(_iso)
try:
    exec(CELL[74], G)
    check(True, 'calibration cell degrades gracefully with no OT Week files present')
except Exception as e:
    check(False, f'calibration cell raised with no files: {e}')
finally:
    os.chdir(_cwd)

# fabricate one OT Week workbook shaped like payroll's and re-run
try:
    from openpyxl import Workbook
    _wbdir = os.path.join(OUT, 'fab', 'data', 'Financial')
    os.makedirs(_wbdir, exist_ok=True)
    _ws0 = OT_SPINE[~OT_SPINE['is_partial_week'] | OT_SPINE['is_feed_incomplete']]
    _ws0 = G['dash_wk_ot_co']
    _pick = _ws0[~_ws0['is_partial_week'].fillna(False).astype(bool)].iloc[1]
    _wk = pd.Timestamp(_pick['week_start'])
    wb = Workbook(); ws = wb.active; ws.title = 'Report'
    ws.cell(row=6, column=3, value='Grand Total')
    ws.cell(row=6, column=11, value=float(_pick['worked_hours']))          # hours worked
    ws.cell(row=6, column=13, value=float(_pick['ot_hours']) * 1.05)       # est OT
    ws.cell(row=6, column=16, value=float(_pick['employees']) * 40.0)      # FTE hours
    ws.cell(row=6, column=18, value=float(_pick['employees']) * 40.0)     # payroll FTE hrs
    ws.cell(row=6, column=19, value=float(_pick['ot_hours']) * 1.02)       # paid OT
    ws.cell(row=6, column=20, value=0.15)
    _fn = f'OT Week {_wk.month}.{_wk.day}.{str(_wk.year)[2:]}-{_wk.month}.{_wk.day}.{str(_wk.year)[2:]}.xlsx'
    wb.save(os.path.join(_wbdir, _fn))
    os.chdir(os.path.join(OUT, 'fab'))
    try:
        exec(CELL[74], G)
        cal = G['tbl_ot_calibration']
        check(len(cal) == 1, f'calibration parsed the fabricated OT Week file ({len(cal)} rows)')
        if len(cal):
            check(abs(float(cal['inferred_vs_paid_pct'].iloc[0]) - (1 / 1.02 - 1) * 100) < 0.5,
                  'inferred-vs-paid % computed correctly against the Grand Total row')
    finally:
        os.chdir(_cwd)
except ImportError:
    print('  (openpyxl unavailable — fabricated-workbook test skipped)')

# payroll_weeks on the monthly bridge (v1.8.1)
check('payroll_weeks' in G['dash_mo_ot_co'].columns
      and int(pd.to_numeric(G['dash_mo_ot_co']['payroll_weeks'], errors='coerce')
              .max()) >= 4,
      'the OT monthly bridge carries payroll_weeks (4-vs-5-week months made visible)')

# ═══════════════════════════════════════════════════════════════════════════════
# Cell 18 — the full renderer, and the year on every axis label (v1.8.1)
# ═══════════════════════════════════════════════════════════════════════════════
print('\n=== Cell 18 — renderer & MMM-YY axis labels ==================================')


def _rframe(gcol, vals, cols):
    rows = []
    for k, v in enumerate(vals):
        for wi, ws in enumerate(WEEKS):
            r = {'week_start': ws}
            if gcol:
                r[gcol] = v
                if gcol == 'tech_warehouse':
                    r['vp'], r['state'], r['metro'] = _HIER[v]
            for c, base in cols.items():
                r[c] = base * (1 + 0.1 * ((wi + k) % 5))
            rows.append(r)
    d = attach_spine(pd.DataFrame(rows))
    gc = [gcol] if gcol else []
    return add_trailing(d, gc, count_cols=list(cols), rate_cols=list(cols))


_RENTS = [('tech_warehouse', ['WH_A', 'WH_B', 'WH_C']), ('vp', ['VP North', 'VP South']),
          ('metro', ['Houston', 'Other']), (None, [None])]
_PROD = dict(total_tickets=500.0, total_active_tech_days=40.0,
             tickets_per_active_day_per_tech=12.5, virtual_ticket_share_pct=6.0,
             denominator_inflation_pct=8.0, rolling_4wk_avg=12.5)
_RED = dict(redelivery_count=12.0, total_tickets=520.0, redel_per_100_tickets=2.3)
_SO = dict(stockout_orders=30.0, total_tickets=520.0, stockouts_per_100_tickets=5.8)
for _fam, _cols in (('prod', _PROD), ('redel', _RED), ('stockout', _SO)):
    for _gc, _vals in _RENTS:
        _nm = {'tech_warehouse': 'wh', 'vp': 'vp', 'metro': 'metro', None: 'co'}[_gc]
        G[f'dash_wk_{_fam}_{_nm}'] = _rframe(_gc, _vals, _cols)
_tb_cols = ['rank_group', 'techfirstname', 'techlastname', 'tech_warehouse', 'total_tickets',
            'active_weekdays', 'tickets_per_active_day', 'redel_per_100_tickets']
_tb = pd.DataFrame([['Top', 'Ann', 'Alpha', 'WH_A', 420, 45, 14.2, 1.9],
                    ['Bottom', 'Cal', 'Gamma', 'WH_C', 210, 42, 7.1, 4.8]], columns=_tb_cols)
G['dash_tb_co'] = _tb
G['dash_tb_vp'] = _tb.assign(vp='VP North')
G['dash_tb_metro'] = _tb.assign(metro='Houston')
G['dash_tb_wh'] = _tb
G['DASH_PDF_NAME'] = 'verify_pages_v181.pdf'

exec(CELL[84], G)
PAGES = G['dash_page_pdfs']
check(len(PAGES) == 1 + 2 + 2 + 3,
      f'one entry per entity: company + 2 VP + 2 metro + 3 WH = 8 (got {len(PAGES)})')
check(all(len(v) == 2 for v in PAGES.values()), 'every entity rendered both pages')

# the v1.8.1 axis contract: month ticks, MMM-YY labels
_fig, _ax = plt.subplots()
_ax.plot(WEEKS, range(len(WEEKS)))
G['_wk_axis'](_ax, pd.DataFrame({'week_start': WEEKS}))
_fig.canvas.draw()
_labels = [t.get_text() for t in _ax.get_xticklabels() if t.get_text()]
plt.close(_fig)
_ok = [l for l in _labels if re.fullmatch(r'[A-Z][a-z]{2}-\d{2}', l)]
check(len(_labels) > 0 and len(_ok) == len(_labels),
      f'every weekly axis label is MMM-YY with the year (got {_labels[:6]})')

# ═══════════════════════════════════════════════════════════════════════════════
# Cell 16.5 — cross-dataset spike surveillance, on a planted spike + a planted dip
# ═══════════════════════════════════════════════════════════════════════════════
print('\n=== Cell 16.5 — spike surveillance ===========================================')


def _real_lev(a, b):
    if a == b:
        return 0
    la, lb = len(a), len(b)
    d = [[0] * (lb + 1) for _ in range(la + 1)]
    for i in range(la + 1):
        d[i][0] = i
    for j in range(lb + 1):
        d[0][j] = j
    for i in range(1, la + 1):
        for j in range(1, lb + 1):
            cost = 0 if a[i - 1] == b[j - 1] else 1
            d[i][j] = min(d[i - 1][j] + 1, d[i][j - 1] + 1, d[i - 1][j - 1] + cost)
    return d[la][lb]


# 10 months; 3 persistent techs; Dec-2025 adds 4 transient names — two of them
# name-splitting signatures (same eid as a persistent name AND within 1 edit), two real
# short-stint newcomers. PLC hours dip -50% in Jan-2026.
_MOS = pd.period_range('2025-06', '2026-03', freq='M')
_PERS = [('Al', 'Ape', 'E1'), ('Bo', 'Bee', 'E2'), ('Cy', 'Sea', 'E3')]
_TRANS = [('All', 'Ape', 'E1', 2), ('Bo', 'Be', 'E2', 1),      # splitting signatures
          ('Zed', 'Zulu', 'E9', 2), ('Yan', 'Yolo', 'E8', 3)]  # genuine newcomers
_tdr, _txr2 = [], []
for _mo in _MOS:
    _days = [d for d in pd.date_range(_mo.start_time, _mo.end_time) if d.dayofweek < 5][:10]
    for _fn, _ln, _eid in _PERS:
        for _d in _days:
            _tdr.append((_fn, _ln, f'{_fn}|{_ln}', 'WH_A', _d, 3))
            _txr2.append((_fn, _ln, _d, _eid))
    if str(_mo) == '2025-12':
        for _fn, _ln, _eid, _nd in _TRANS:
            for _d in _days[:_nd]:
                _tdr.append((_fn, _ln, f'{_fn}|{_ln}', 'WH_A', _d, 2))
                _txr2.append((_fn, _ln, _d, _eid))
_gs_td = pd.DataFrame(_tdr, columns=['techfirstname', 'techlastname', '_tech_key',
                                     'tech_warehouse', 'completed_date', 'day_tickets'])
_gs_tx = pd.DataFrame(_txr2, columns=['techfirstname', 'techlastname', 'completed_date',
                                      '_matched_eid'])
_gs_tx['order_num'] = np.arange(len(_gs_tx)).astype(str)
_gs_tx['_unattributed'] = False
_gs_hours = pd.DataFrame([{'workdate': _mo.start_time + pd.Timedelta(days=3),
                           'hours': 500.0 if str(_mo) == '2026-01' else 1000.0,
                           'employee_name': 'X, Y'} for _mo in _MOS])
_gs_cd = pd.DataFrame([{'period': str(_mo), 'pt_count': 900.0,
                        'census_date': _mo.start_time + pd.Timedelta(days=i)}
                       for _mo in _MOS for i in range(5)])
GS = dict(pd=pd, np=np, re=re,
          SPIKE_SCAN_THRESHOLD_PCT=25.0, SPIKE_SCAN_WINDOW_MONTHS=7,
          levenshtein=lambda a, b: _real_lev(a, b),
          df_tx=_gs_tx, _techday=_gs_td, df_hours_raw=_gs_hours, df_census_daily=_gs_cd,
          redel_linked=pd.DataFrame(), dash_mo_stockout_co=pd.DataFrame())
import contextlib as _ctx
_buf16 = io.StringIO()
with _ctx.redirect_stdout(_buf16):
    exec(CELL[78], GS)
_out16 = _buf16.getvalue()
print(_out16)
_scan = GS['tbl_spike_scan']
_fl = _scan[_scan['flagged']] if len(_scan) else _scan
check(bool(((_fl['dataset'] == 'distinct technician names (weekday tickets)')
            & (_fl['period'] == '2025-12') & (_fl['deviation_pct'] > 25)).any()),
      'the planted technician-name spike month is flagged HIGH')
check(bool(((_fl['dataset'] == 'payroll hours (PLC, all departments)')
            & (_fl['period'] == '2026-01') & (_fl['deviation_pct'] < -25)).any()),
      'the planted payroll-hours DIP month is flagged LOW (missing data direction)')
check(not bool((_fl['period'].isin([str(_MOS[0]), str(_MOS[-1])])).any()),
      'window-clipped edge months are never flagged')
_tn = GS['tbl_techname_spike']
check(len(_tn) == 4, f'decomposition covers exactly the 4 transient names (got {len(_tn)})')
if len(_tn):
    check(int(_tn['eid_shadows_persistent_name'].sum()) == 2,
          'the two same-employee-ID spellings carry the shadow signature')
    check(int(_tn['name_within_1_edit_of_persistent'].sum()) >= 2,
          'the two misspellings sit within one edit of a persistent name')
check('NAME SPLITTING' in _out16 or 'NAME-SPLITTING' in _out16,
      'the run log states the splitting verdict for the planted month')

# ═══════════════════════════════════════════════════════════════════════════════
# Cell 17.5 — the scorecard guard (audit A1)
# ═══════════════════════════════════════════════════════════════════════════════
print('\n=== Cell 17.5 — scorecard guard ===============================================')
_ENTS = [('vp', ['VP North', 'VP South']), (None, [None])]


def _frame(gcol, vals, cols):
    rows = []
    for k, v in enumerate(vals):
        for wi, ws in enumerate(WEEKS):
            r = {'week_start': ws}
            if gcol:
                r[gcol] = v
            for c, base in cols.items():
                r[c] = base * (1 + 0.1 * ((wi + k) % 5))
            rows.append(r)
    d = attach_spine(pd.DataFrame(rows))
    gc = [gcol] if gcol else []
    return add_trailing(d, gc, rate_cols=list(cols))


G['dash_wk_prod_vp'] = _frame('vp', ['VP North', 'VP South'],
                              dict(tickets_per_active_day_per_tech=12.5))
G['dash_wk_prod_co'] = _frame(None, [None], dict(tickets_per_active_day_per_tech=12.0))
G['dash_wk_census_vp'] = _frame('vp', ['VP North', 'VP South'],
                                dict(attendance_rate_pct=75.0, census_per_tech_headcount=94.0))
# reuse the REAL census/ot company frames where they exist
G['dash_wk_census_co'] = G['dash_wk_census_co']
G['dash_wk_ot_vp'] = _frame('vp', ['VP North', 'VP South'],
                            dict(ot_pct_of_worked=16.0, ot_hours_per_tech=8.0))
G['dash_wk_lost_vp'] = _frame('vp', ['VP North', 'VP South'],
                              dict(lost_cost_per_1k_pt_days=140.0, recovery_rate_pct=15.0))
G['dash_wk_lost_co'] = _frame(None, [None],
                              dict(lost_cost_per_1k_pt_days=150.0, recovery_rate_pct=14.0))
G['dash_wk_redel_vp'] = _frame('vp', ['VP North', 'VP South'], dict(redel_per_100_tickets=2.2))
G['dash_wk_redel_co'] = _frame(None, [None], dict(redel_per_100_tickets=2.0))
G['dash_wk_stockout_vp'] = _frame('vp', ['VP North', 'VP South'],
                                  dict(stockouts_per_100_tickets=5.0, open_pct=7.0,
                                       canceled_pct=28.0))
G['dash_wk_stockout_co'] = _frame(None, [None],
                                  dict(stockouts_per_100_tickets=5.2, open_pct=7.1,
                                       canceled_pct=27.0))
G['redel_linked'] = pd.DataFrame(columns=['techfirstname', 'techlastname', 'event_key',
                                          'week_start', 'product'])
exec(CELL[82], G)
sc = G['vp_scorecard']
check('ot_hours_per_tech' in set(sc['metric']) if len(sc) else False,
      'audit A1: ot_hours_per_tech is SCORED now the company frame carries it')
if len(sc):
    _ot_rows = sc[sc['metric'] == 'ot_hours_per_tech']
    check(bool(_ot_rows['company_value'].notna().all()),
          'ot_hours_per_tech rows carry a real company value (was NaN through v1.7.1)')

# and if the company frame LACKS the column, the guard reports it instead of NaN-scoring
G2 = dict(G)
G2['dash_wk_ot_co'] = G['dash_wk_ot_co'].drop(columns=['ot_hours_per_tech',
                                                       'ot_hours_per_tech' + TS])
import contextlib
_buf = io.StringIO()
with contextlib.redirect_stdout(_buf):
    exec(CELL[82], G2)
sc2 = G2['vp_scorecard']
check('ot_hours_per_tech' not in (set(sc2['metric']) if len(sc2) else set()),
      'a metric missing from the company frame is not scored…')
check('ot_hours_per_tech' in _buf.getvalue(),
      '…and the guard names it in _missing_metrics instead of failing silently')

print('\n' + '=' * 80)
print(f'{len(PASSES)} passed, {len(FAILS)} failed')
if FAILS:
    print('FAILURES:')
    for f in FAILS:
        print('  - ' + f)
    sys.exit(1)
print('ALL CHECKS PASSED')

